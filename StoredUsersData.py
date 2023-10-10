import sqlite3
from openpyxl import load_workbook

# Load the Excel workbook
workbook = load_workbook('user_data.xlsx')
sheet = workbook.active

# Connect to the SQLite database
conn = sqlite3.connect('Store.db')
cursor = conn.cursor()

# Drop the Users table if it exists
cursor.execute('DROP TABLE IF EXISTS Users')

# Create the Users table
cursor.execute('''
    CREATE TABLE Users (
        UserID INTEGER PRIMARY KEY,
        UserName TEXT,
        Password TEXT,
        FullName TEXT,
        Address TEXT,
        PhoneNumber TEXT,
        Email TEXT
    )
''')

# Insert data from the Excel sheet into the Users table
for row in sheet.iter_rows(min_row=2, values_only=True):
    user_id, user_name, password, full_name, address, phone_number, email = row
    cursor.execute('''
        INSERT INTO Users (UserID, UserName, Password, FullName, Address, PhoneNumber, Email)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (user_id, user_name, password, full_name, address, phone_number, email))

# Commit the changes and close the database connection
conn.commit()
conn.close()

print("Data saved to Store.db successfully!")
