import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import random
import requests
import sqlite3

# Connect to SQLite database
conn = sqlite3.connect("Store.db")
cursor = conn.cursor()

# Drop Products table if it exists
cursor.execute("DROP TABLE IF EXISTS Products")

# Create Products table
cursor.execute("CREATE TABLE Products (ProductID INTEGER PRIMARY KEY, ProductName TEXT, Price REAL, Description TEXT, Quantity INTEGER)")

# Create a new workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Define the number of products to create
num_products = 2000

# Define the headers for the product data
headers = ['Product ID', 'Name', 'Price', 'Description', 'Quantity']

# Write the headers to the first row of the sheet
for col_num, header in enumerate(headers, start=1):
    column_letter = get_column_letter(col_num)
    sheet[column_letter + '1'] = header
    sheet[column_letter + '1'].font = Font(bold=True)

# Fetch product names and descriptions from the Fakestore API
response = requests.get('https://fakestoreapi.com/products')
products = response.json()

# Generate and write the product data to the sheet
for product_id in range(1, num_products + 1):
    product = random.choice(products)
    name = product['title']
    price = round(product['price'], 2)
    description = product['description']
    quantity = random.randint(1, 100)

    sheet.cell(row=product_id + 1, column=1, value=product_id)
    sheet.cell(row=product_id + 1, column=2, value=name)
    sheet.cell(row=product_id + 1, column=3, value=price)
    sheet.cell(row=product_id + 1, column=4, value=description)
    sheet.cell(row=product_id + 1, column=5, value=quantity)

# Save the workbook to a file
workbook.save('products.xlsx')

# Read the product data from the Excel file
excel_file = "products.xlsx"
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

# Insert data into the table
for row in sheet.iter_rows(min_row=2, values_only=True):
    product_id, name, price, description, quantity = row
    cursor.execute("INSERT INTO Products (ProductID, ProductName, Price, Description, Quantity) VALUES (?, ?, ?, ?, ?)",
                   (product_id, name, price, description, quantity))

# Commit the changes and close the connection
conn.commit()
conn.close()

print("Data successfully stored in the database.")
