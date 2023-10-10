import openpyxl
import random

# List of realistic usernames
usernames = ['john_doe', 'mary_smith', 'alex_jones', 'sara_wilson', 'michael_brown', 'laura_johnson']

# List of realistic surnames
surnames = ['Smith', 'Johnson', 'Williams', 'Jones', 'Brown', 'Davis', 'Miller', 'Wilson', 'Moore', 'Taylor']

# Create a new workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Write the column headers
sheet['A1'] = 'UserID'
sheet['B1'] = 'UserName'
sheet['C1'] = 'Password'
sheet['D1'] = 'Full Name'
sheet['E1'] = 'Address'
sheet['F1'] = 'Phone Number'
sheet['G1'] = 'Email'

# Generate data for 1000 rows
for i in range(1, 1001):
    user_id = i
    user_name = random.choice(usernames)
    password = 'password' + str(i).zfill(4)  # Padding user ID with zeros to ensure 8 characters
    full_name = user_name + ' ' + random.choice(surnames)
    address = 'Lubbock ' + str(i)
    phone_number = '123-45-789' + str(i).zfill(2)  # Padding phone number with zeros to ensure 8 characters
    email = 'user' + str(i).zfill(4) + '@example.com'  # Padding user ID with zeros to ensure 8 characters

    # Write the data to the sheet
    sheet.cell(row=i + 1, column=1, value=user_id)
    sheet.cell(row=i + 1, column=2, value=user_name)
    sheet.cell(row=i + 1, column=3, value=password)
    sheet.cell(row=i + 1, column=4, value=full_name)
    sheet.cell(row=i + 1, column=5, value=address)
    sheet.cell(row=i + 1, column=6, value=phone_number)
    sheet.cell(row=i + 1, column=7, value=email)

# Save the workbook
workbook.save('user_data.xlsx')
