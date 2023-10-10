import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import random

def fetch_product_data():
    response = requests.get('https://fakestoreapi.com/products')
    products = response.json()
    return products

def generate_orders_excel():
    # Fetch product data from the API
    products = fetch_product_data()

    # Create a new workbook
    workbook = Workbook()
    sheet = workbook.active

    # Set column headers
    headers = ['OrderID', 'Product Name', 'Quantity', 'Total Price']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header

    # Generate orders and populate the sheet with data
    order_count = 10000
    order_id_prefix = 'FK'
    min_products_per_order = 10

    for order_index in range(1, order_count + 1):
        order_id = f'{order_id_prefix}{order_index:08d}'
        num_products = random.randint(min_products_per_order, len(products))
        order_products = random.sample(products, num_products)

        for product in order_products:
            product_name = product['title']
            quantity = random.randint(1, 10)
            price = float(product['price'])
            total_price = quantity * price

            sheet.append([order_id, product_name, quantity, total_price])

    # Save the workbook as "orders.xlsx"
    workbook.save('orders.xlsx')

    print("orders.xlsx generated successfully!")

# Call the function to generate the Excel file
generate_orders_excel()
